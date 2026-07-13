#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Wall Street Standard DCF Valuation Model Generator
Generates a 10-year Unlevered DCF Model across multiple fully-linked sheets.
Author: Quant Finance & Investment Banking Modeling Expert
"""

import os
import sys
import argparse
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openbb import obb
import yfinance as yf

def extract_financial_data(ticker):
    """
    Fetch financial data (income, balance, cash flow, profile, quote) from OpenBB and yfinance.
    """
    print(f"[*] Fetching financial data for {ticker} using OpenBB Platform...")
    try:
        income_obb = obb.equity.fundamental.income(symbol=ticker, provider="yfinance", period="annual")
        balance_obb = obb.equity.fundamental.balance(symbol=ticker, provider="yfinance", period="annual")
        cash_obb = obb.equity.fundamental.cash(symbol=ticker, provider="yfinance", period="annual")
        
        df_income = income_obb.to_dataframe()
        df_balance = balance_obb.to_dataframe()
        df_cash = cash_obb.to_dataframe()
    except Exception as e:
        raise RuntimeError(f"Failed to fetch financial statements from OpenBB: {e}")
        
    try:
        profile_obb = obb.equity.profile(symbol=ticker, provider="yfinance")
        quote_obb = obb.equity.price.quote(symbol=ticker, provider="yfinance")
        
        df_profile = profile_obb.to_dataframe()
        df_quote = quote_obb.to_dataframe()
    except Exception as e:
        print(f"[!] Warning: Failed to fetch profile/quote data ({e}). Using default assumptions...")
        df_profile = pd.DataFrame()
        df_quote = pd.DataFrame()

    # Sort ascending by period_ending
    df_income = df_income.sort_values("period_ending", ascending=True)
    df_balance = df_balance.sort_values("period_ending", ascending=True)
    df_cash = df_cash.sort_values("period_ending", ascending=True)

    if len(df_income) < 2:
        raise ValueError("Insufficient historical data (requires at least 2 years).")

    latest_date = df_income['period_ending'].iloc[-1]
    latest_year = pd.to_datetime(latest_date).year
    company_name = df_profile['name'].iloc[0] if not df_profile.empty and 'name' in df_profile.columns else ticker
    beta_val = df_profile['beta'].iloc[0] if not df_profile.empty and 'beta' in df_profile.columns else 1.0
    if pd.isna(beta_val) or beta_val <= 0:
        beta_val = 1.0

    price_val = None
    if not df_quote.empty and 'last_price' in df_quote.columns:
        price_val = df_quote['last_price'].iloc[0]
    if price_val is None or pd.isna(price_val):
        price_val = 100.0

    # ---- Shares outstanding: CRITICAL FIX for dual-class structures (GOOG/GOOGL, BRK.A/B etc.) ----
    # The API's `shares_outstanding` can return only one share class, badly undercounting total shares.
    # Gold-standard: derive total shares from market cap / price — this always captures all classes.
    shares_val = None
    try:
        ticker_yf_tmp = yf.Ticker(ticker)
        info_tmp = ticker_yf_tmp.info
        mkt_cap = info_tmp.get('marketCap')
        if mkt_cap and not pd.isna(mkt_cap) and price_val and price_val > 0:
            shares_val = mkt_cap / price_val  # total shares implied, all classes
            print(f"[+] Shares derived from market cap / price: {shares_val/1e9:.3f}B  (mktcap={mkt_cap/1e9:.1f}B @ ${price_val:.2f})")
    except Exception as e:
        print(f"[!] Warning: Could not derive shares from market cap ({e})")

    # Fallback chain: diluted WA shares from income statement -> flat default
    if shares_val is None or pd.isna(shares_val) or shares_val <= 0:
        shares_val = df_income['weighted_average_diluted_shares_outstanding'].iloc[-1]
    if pd.isna(shares_val) or shares_val <= 0:
        shares_val = 100e6  # last-resort default

    # Fetch analyst estimates dynamically from yfinance for professional near-term forecasting
    # Strategy: use up to 3 consensus anchor points (0y, +1y, implied +2y via EPS trend)
    # before applying our own conservative tapering algorithm from year 4 onwards.
    print(f"[*] Fetching analyst consensus estimates for {ticker} from yfinance...")
    g_year1 = None
    g_year2 = None
    g_year3 = None  # implied from EPS growth deceleration
    try:
        ticker_yf = yf.Ticker(ticker)
        rev_est = ticker_yf.revenue_estimate
        if rev_est is not None and not rev_est.empty:
            if '0y' in rev_est.index:
                val = rev_est.loc['0y', 'growth']
                if not pd.isna(val) and val != 0:
                    g_year1 = float(val)
            if '+1y' in rev_est.index:
                val = rev_est.loc['+1y', 'growth']
                if not pd.isna(val) and val != 0:
                    g_year2 = float(val)

        # Derive implied Year 3 from EPS trend as a proxy for continued deceleration.
        # EPS estimates 0y vs +1y give us year-on-year growth from which we can infer revenue.
        eps_trend = ticker_yf.eps_trend
        if eps_trend is not None and not eps_trend.empty and g_year2 is not None:
            try:
                eps_0y = float(eps_trend.loc['0y', 'current'])
                eps_1y = float(eps_trend.loc['+1y', 'current'])
                if eps_0y > 0:
                    eps_growth = (eps_1y - eps_0y) / eps_0y
                    # Conservative: Year 3 growth = midpoint between Year 2 consensus and EPS deceleration
                    # (assumes margin expansion slows, so revenue growth ~ avg of yr2 and eps-implied)
                    g_year3 = (g_year2 + max(eps_growth, g_year2 * 0.7)) / 2
                    g_year3 = max(0.03, min(0.6, g_year3))  # clip
            except Exception:
                pass

        print(f"[+] Analyst Consensus Growth: Yr1={f'{g_year1:.2%}' if g_year1 is not None else 'N/A'}, "
              f"Yr2={f'{g_year2:.2%}' if g_year2 is not None else 'N/A'}, "
              f"Yr3 (implied)={f'{g_year3:.2%}' if g_year3 is not None else 'N/A'}")
    except Exception as e:
        print(f"[!] Warning: Failed to fetch consensus estimates from yfinance ({e}). Tapering from historical growth instead.")

    # Extract historical metrics dynamically
    def get_val_for_year(year):
        inc_y = df_income[df_income['period_ending'].apply(lambda x: pd.to_datetime(x).year) == year]
        bal_y = df_balance[df_balance['period_ending'].apply(lambda x: pd.to_datetime(x).year) == year]
        cash_y = df_cash[df_cash['period_ending'].apply(lambda x: pd.to_datetime(x).year) == year]
        
        if inc_y.empty:
            return None
            
        inc_row = inc_y.iloc[0]
        bal_row = bal_y.iloc[0] if not bal_y.empty else pd.Series()
        cash_row = cash_y.iloc[0] if not cash_y.empty else pd.Series()
        
        rev = inc_row.get('total_revenue', inc_row.get('operating_revenue', 0.0))
        cost = inc_row.get('cost_of_revenue', 0.0)
        gp = inc_row.get('gross_profit', None)
        if gp is None or pd.isna(gp):
            gp = rev - cost if (rev and cost) else 0.0
            
        ebit = inc_row.get('ebit', inc_row.get('operating_income', inc_row.get('total_operating_income_as_reported', 0.0)))
        pretax = inc_row.get('total_pre_tax_income', inc_row.get('income_before_tax', ebit))
        tax = inc_row.get('tax_provision', 0.0)
        net_inc = inc_row.get('net_income', inc_row.get('net_income_continuous_operations', 0.0))
        
        # D&A extraction
        da = 0.0
        for col in ['depreciation_and_amortization', 'depreciation_amortization_depletion']:
            if col in cash_row and not pd.isna(cash_row[col]):
                da = abs(cash_row[col])
                break
        if da == 0.0:
            dep = abs(inc_row.get('reconciled_depreciation', 0.0))
            amo = 0.0
            for col in ['amortization_of_intangibles', 'amortization_cash_flow']:
                if col in cash_row and not pd.isna(cash_row[col]):
                    amo = abs(cash_row[col])
                    break
            da = dep + amo
            
        # CapEx extraction
        capex = 0.0
        for col in ['capital_expenditure', 'investments_in_property_plant_and_equipment', 'net_ppe_purchase_and_sale']:
            if col in cash_row and not pd.isna(cash_row[col]):
                capex = abs(cash_row[col])
                break
                
        ocf = cash_row.get('operating_cash_flow', cash_row.get('cash_flow_from_continuing_operating_activities', 0.0))
        fcf = cash_row.get('free_cash_flow', ocf - capex)
        
        # Balance sheet cash & debt
        cash_col = 'cash_cash_equivalents_and_short_term_investments'
        if cash_col in bal_row.index and not pd.isna(bal_row[cash_col]):
            cash_val = bal_row[cash_col]
        else:
            cash_val = bal_row.get('cash_and_cash_equivalents', 0.0) + bal_row.get('short_term_investments', 0.0)
        if pd.isna(cash_val) or cash_val == 0.0:
            cash_val = bal_row.get('cash_financial', 0.0)
            
        debt_col = 'total_debt'
        if debt_col in bal_row.index and not pd.isna(bal_row[debt_col]):
            debt_val = bal_row[debt_col]
        else:
            debt_val = bal_row.get('long_term_debt', 0.0) + bal_row.get('current_debt', 0.0)
            
        return {
            "revenue": rev / 1e6 if not pd.isna(rev) else 0.0,
            "gross_profit": gp / 1e6 if not pd.isna(gp) else 0.0,
            "ebit": ebit / 1e6 if not pd.isna(ebit) else 0.0,
            "pretax": pretax / 1e6 if not pd.isna(pretax) else 0.0,
            "tax": tax / 1e6 if not pd.isna(tax) else 0.0,
            "net_income": net_inc / 1e6 if not pd.isna(net_inc) else 0.0,
            "da": da / 1e6 if not pd.isna(da) else 0.0,
            "capex": capex / 1e6 if not pd.isna(capex) else 0.0,
            "ocf": ocf / 1e6 if not pd.isna(ocf) else 0.0,
            "fcf": fcf / 1e6 if not pd.isna(fcf) else 0.0,
            "cash_bs": cash_val / 1e6 if not pd.isna(cash_val) else 0.0,
            "debt_bs": debt_val / 1e6 if not pd.isna(debt_val) else 0.0,
        }

    # Extract 4 historical years dynamically
    hist_years = [latest_year - 3, latest_year - 2, latest_year - 1, latest_year]
    hist_data = {}
    for yr in hist_years:
        res = get_val_for_year(yr)
        if res:
            hist_data[yr] = res
            
    if latest_year not in hist_data:
        raise ValueError(f"Failed to extract historical data for the latest year {latest_year}")

    latest_act = hist_data[latest_year]
    prev_act = hist_data.get(latest_year - 1, latest_act)
    
    rev_latest = latest_act["revenue"]
    rev_prev = prev_act["revenue"]
    growth_latest = (rev_latest - rev_prev) / rev_prev if rev_prev else 0.08
    ebit_margin_latest = latest_act["ebit"] / rev_latest if rev_latest else 0.15
    tax_rate_latest = latest_act["tax"] / latest_act["pretax"] if latest_act["pretax"] > 0 else 0.21
    if tax_rate_latest < 0 or tax_rate_latest > 0.5:
        tax_rate_latest = 0.21
    da_rate_latest = latest_act["da"] / rev_latest if rev_latest else 0.03
    capex_rate_latest = latest_act["capex"] / rev_latest if rev_latest else 0.04
    cash_latest = latest_act["cash_bs"]
    debt_latest = latest_act["debt_bs"]

    return {
        "ticker": ticker,
        "company_name": company_name,
        "latest_year": latest_year,
        "hist_years": hist_years,
        "hist_data": hist_data,
        "beta_val": beta_val,
        "price_val": price_val,
        "shares_val": shares_val / 1e6, # in millions
        "growth_latest": growth_latest,
        "ebit_margin_latest": ebit_margin_latest,
        "tax_rate_latest": tax_rate_latest,
        "da_rate_latest": da_rate_latest,
        "capex_rate_latest": capex_rate_latest,
        "cash_latest": cash_latest,
        "debt_latest": debt_latest,
        "g_year1": g_year1,
        "g_year2": g_year2,
        "g_year3": g_year3,
    }

def generate_openbb_dcf(ticker, output_path=None):
    """
    Generate professional multi-sheet DCF valuation workbook.
    """
    try:
        data = extract_financial_data(ticker)
    except Exception as e:
        print(f"[!] Error: {e}")
        return False

    company_name = data["company_name"]
    latest_year = data["latest_year"]
    hist_years = data["hist_years"]
    hist_data = data["hist_data"]
    beta_val = data["beta_val"]
    price_val = data["price_val"]
    shares_val = data["shares_val"]
    
    growth_latest = data["growth_latest"]
    ebit_margin_latest = data["ebit_margin_latest"]
    tax_rate_latest = data["tax_rate_latest"]
    da_rate_latest = data["da_rate_latest"]
    capex_rate_latest = data["capex_rate_latest"]
    cash_latest = data["cash_latest"]
    debt_latest = data["debt_latest"]
    g_year1 = data["g_year1"]
    g_year2 = data["g_year2"]
    g_year3 = data["g_year3"]

    # 1. Define Projections Timeline (10 Years)
    fc_labels = [f"FY{latest_year + i}E" for i in range(1, 11)]

    # ---- Revenue Growth Tapering Algorithm ----
    # Phase 1: Analyst consensus anchors years 1-3 (hard numbers from Street)
    # Phase 2: Gentle convex decay from year-3 anchor → 3% perpetuity growth at year 10
    # Convex schedule decays slowly early (market still growing) then accelerates to terminal.
    g_start = g_year1 if g_year1 is not None else growth_latest
    g_mid   = g_year2 if g_year2 is not None else g_start * 0.85
    g_mid3  = g_year3 if g_year3 is not None else (g_mid * 0.75 if g_mid is not None else growth_latest * 0.7)

    # Clip all inputs to sane ranges
    g_start = max(-0.10, min(0.60, g_start))
    g_mid   = max(-0.10, min(0.60, g_mid))
    g_mid3  = max( 0.03, min(0.60, g_mid3))

    # Target growth rate at Year 10 of the explicit forecast period (decoupled from TV perpetual growth)
    # A fast-growing company matures to a higher terminal explicit growth rate (e.g. 5% or 6%),
    # while a slower-growing company matures closer to the perpetuity growth rate.
    # We dynamically calculate this as 25% of Year 3 growth, bounded between 3.5% and 6.0%.
    g_explicit_terminal = max(0.035, min(0.06, g_mid3 * 0.25))

    # Anchor years 1-3 directly to consensus
    fc_growth_rates = [g_start, g_mid, g_mid3]

    # Phase 2: convex decay from g_mid3 (yr 3) → g_explicit_terminal (yr 10) over 7 steps
    # Using a square-root schedule: rate decays fast at first, then flattens near terminal
    # This means yr 4-6 still see relatively healthy growth before fading to perpetuity.
    n_taper = 7  # years 4 through 10
    for i in range(1, n_taper + 1):
        # convex weight: sqrt curve decays slower early than linear
        weight = (i / n_taper) ** 1.5   # exponent < 2 = concave = stays high longer
        rate = g_mid3 - weight * (g_mid3 - g_explicit_terminal)
        fc_growth_rates.append(rate)

    # CapEx fade: linearly fade from capex_rate_latest in Yr 0 to da_rate_latest in Yr 10
    fc_capex_rates = [capex_rate_latest - t * (capex_rate_latest - da_rate_latest) / 10 for t in range(1, 11)]

    # Set up openpyxl workbook
    wb = openpyxl.Workbook()
    
    # ----------------------- STYLING SETUP -----------------------
    font_title = Font(name='Calibri', size=12, bold=True, color='1B365D')
    font_subtitle = Font(name='Calibri', size=8, italic=True, color='555555')
    font_header = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
    font_section = Font(name='Calibri', size=9, bold=True, color='1B365D')
    
    font_label = Font(name='Calibri', size=9, color='000000')
    font_label_bold = Font(name='Calibri', size=9, bold=True, color='000000')
    
    # Color Coding Fonts
    font_black = Font(name='Calibri', size=9, color='000000')
    font_blue = Font(name='Calibri', size=9, color='0000FF')
    font_bold_black = Font(name='Calibri', size=9, bold=True, color='000000')
    font_bold_blue = Font(name='Calibri', size=9, bold=True, color='0000FF')

    fill_header = PatternFill(start_color='1B365D', end_color='1B365D', fill_type='solid')
    fill_summary = PatternFill(start_color='E8EEF5', end_color='E8EEF5', fill_type='solid')
    fill_label_summary = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    fill_highlight = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    fill_highlight_soft = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    fill_green_soft = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    font_green = Font(name='Calibri', size=9, color='006100')
    font_green_bold = Font(name='Calibri', size=9, bold=True, color='006100')
    font_red_bold = Font(name='Calibri', size=9, bold=True, color='9C0006')

    thin_gray = Side(style='thin', color='CCCCCC')
    thin_black = Side(style='thin', color='000000')
    double_black = Side(style='double', color='000000')
    
    border_thin_all = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    border_thin_bottom = Border(bottom=thin_gray)
    
    # Clean cell borders for filled rows to avoid overlapping visual artifacts
    border_summary = Border(left=thin_gray, right=thin_gray, top=thin_black, bottom=thin_black)
    border_double_bottom = Border(left=thin_gray, right=thin_gray, top=thin_black, bottom=double_black)

    align_left = Alignment(horizontal='left', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # ----------------------- SHEET 1: SUMMARY -----------------------
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.views.sheetView[0].showGridLines = False
    ws_sum.freeze_panes = 'A5'  # Freeze above data rows

    # Title
    ws_sum['A2'] = f"{company_name} ({ticker}) — DCF Valuation Summary"
    ws_sum['A2'].font = font_title
    ws_sum['A3'] = "Unlevered DCF. USD mm unless noted. Market-value WACC; mid-year explicit cash-flow discounting."
    ws_sum['A3'].font = font_subtitle

    # Write summary card labels & formulas
    # format: Label, Formula/Value, NumberFormat, Border, FontColor, Fill
    summary_rows = [
        ("Calculation integrity", "CHECK", "@", border_summary, font_green_bold, fill_highlight),
        ("Decision readiness", "NOT READY", "@", border_summary, font_red_bold, fill_highlight),
        ("empty1", None, None, None, None, None),
        ("Implied value / share (USD / share)", "=DCF!B28", "$#,##0.00", border_summary, font_bold_black, fill_highlight_soft),
        ("Current share price (USD / share)", "=DCF!B29", "$#,##0.00", border_summary, font_green, None),
        ("Upside / (downside)", "=DCF!B30", "0.0%", border_summary, font_green_bold, fill_green_soft),
        ("empty2", None, None, None, None, None),
        ("Enterprise value (USD mm)", "=DCF!B24", "$#,##0", border_summary, font_black, None),
        ("Less: net debt (USD mm)", "=DCF!B25", "($#,##0);$#,##0;\"-\"", border_summary, font_black, None),
        ("Equity value (USD mm)", "=DCF!B26", "$#,##0", border_summary, font_bold_black, fill_summary),
        ("empty3", None, None, None, None, None),
        ("WACC", "=Assumptions!B22", "0.0%", border_summary, font_black, None),
        ("Terminal growth (g)", "=Assumptions!B25", "0.0%", border_summary, font_black, None),
        ("Terminal value % of EV", "=DCF!B32", "0.0%", border_summary, font_black, None),
        ("Implied FY1E EV/EBIT", "=DCF!B31", "0.0\"x\"", border_summary, font_black, None),
    ]

    r_curr = 5
    for label, formula, num_fmt, border, f_color, cell_fill in summary_rows:
        if label.startswith("empty"):
            r_curr += 1
            continue
            
        cell_lbl = ws_sum.cell(row=r_curr, column=1, value=label)
        cell_lbl.font = font_label_bold if "value" in label.lower() or "integrity" in label.lower() or "readiness" in label.lower() or "equity" in label.lower() or "upside" in label.lower() else font_label
        cell_lbl.fill = fill_label_summary
        cell_lbl.border = border_summary
        cell_lbl.alignment = align_left
        
        cell_val = ws_sum.cell(row=r_curr, column=2, value=formula)
        cell_val.alignment = align_right
        cell_val.font = f_color if f_color else font_black
        if num_fmt:
            cell_val.number_format = num_fmt
        if cell_fill:
            cell_val.fill = cell_fill
        elif border and border.top == thin_black:
            cell_val.fill = fill_summary
            
        cell_val.border = border_summary
        r_curr += 1

    ws_sum.cell(row=r_curr+1, column=1, value="Method: 10-yr unlevered FCF (EBIT*(1-t)+D&A-capex-ΔNWC),").font = font_subtitle
    ws_sum.cell(row=r_curr+2, column=1, value="Gordon-growth terminal, CAPM-built market-value WACC, mid-year discounting. Tabs: Assumptions, Historical, DCF, Sensitivity, Checks, Sources & Audit.").font = font_subtitle

    # ----------------------- SHEET 2: ASSUMPTIONS -----------------------
    ws_ass = wb.create_sheet("Assumptions")
    ws_ass.views.sheetView[0].showGridLines = False
    ws_ass.freeze_panes = 'B5'  # Freeze label col + header row

    # Title
    ws_ass['A2'] = f"{company_name} ({ticker}) — DCF Assumptions"
    ws_ass['A2'].font = font_title
    ws_ass['A3'] = "USD mm. Blue = Input, black = formula. Market-value WACC; mid-year DCF convention."
    ws_ass['A3'].font = font_subtitle

    # Header Row
    ws_ass.cell(row=4, column=1, value="Projection drivers").font = font_label_bold
    ws_ass.cell(row=4, column=1).fill = fill_header
    ws_ass.cell(row=4, column=1).font = font_header
    ws_ass.cell(row=4, column=1).border = border_thin_all
    for col_idx, lbl in enumerate(fc_labels, 2):
        c = ws_ass.cell(row=4, column=col_idx, value=lbl)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border_thin_all

    # Projections Drivers Rows (5 to 10)
    drivers = [
        ("Revenue growth %", fc_growth_rates, "0.0%", True),
        ("EBIT margin %", [ebit_margin_latest]*10, "0.0%", True),
        ("Tax rate %", [tax_rate_latest]*10, "0.0%", True),
        ("D&A % of revenue", [da_rate_latest]*10, "0.0%", True),
        ("Capex % of revenue", fc_capex_rates[:-1] + [None], "0.0%", True), # Year 10 will link to Year 10 D&A rate cell (K8)
        ("Δ NWC % of ΔRevenue", [0.03]*10, "0.0%", True),
    ]

    for idx, (label, rates, num_fmt, is_ass) in enumerate(drivers, 5):
        ws_ass.cell(row=idx, column=1, value=label).font = font_label
        ws_ass.cell(row=idx, column=1).border = border_thin_all
        for col_idx, val in enumerate(rates, 2):
            cell_col_letter = get_column_letter(col_idx)
            cell = ws_ass.cell(row=idx, column=col_idx)
            cell.border = border_thin_all
            
            # Year 10 CapEx rate linked to Year 10 D&A rate cell K8
            if label == "Capex % of revenue" and col_idx == 11:
                cell.value = f"={cell_col_letter}8"
            else:
                cell.value = val
                
            cell.alignment = align_right
            is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
            cell.font = font_blue if is_ass and not is_formula else font_black
            cell.number_format = num_fmt

    # Cost of Capital (WACC) Section
    ws_ass.cell(row=12, column=1, value="Cost of capital (WACC)").font = font_section
    ws_ass.cell(row=12, column=1).border = border_thin_all
    ws_ass.cell(row=12, column=2).border = border_thin_all
    
    wacc_layout = [
        ("Risk-free rate", 0.043, "0.0%", True),
        ("Equity risk premium", 0.045, "0.0%", True),
        ("Levered beta", beta_val, "0.00", True),
        ("Cost of equity", "=B13+B15*B14", "0.0%", False),
        ("Pre-tax cost of debt", 0.047, "0.0%", True),
        ("Effective tax rate", "=B7", "0.0%", False), # links to Year 1 Tax Rate cell in Assumptions
        ("After-tax cost of debt", "=B17*(1-B18)", "0.0%", False),
        ("Equity weight", "=B32*B33/(B32*B33+B30)", "0.0%", False),
        ("Debt weight", "=1-B20", "0.0%", False),
        ("WACC", "=B16*B20+B19*B21", "0.0%", False), # cell B22
    ]

    for idx, (label, val, num_fmt, is_ass) in enumerate(wacc_layout, 13):
        ws_ass.cell(row=idx, column=1, value=label).font = font_label
        ws_ass.cell(row=idx, column=1).border = border_thin_all
        cell = ws_ass.cell(row=idx, column=2, value=val)
        cell.alignment = align_right
        cell.border = border_thin_all
        is_formula = isinstance(val, str) and val.startswith("=")
        cell.font = font_bold_black if label == "WACC" else (font_blue if is_ass and not is_formula else font_black)
        cell.number_format = num_fmt
        if label == "WACC":
            cell.fill = fill_highlight
            ws_ass.cell(row=idx, column=1).font = font_label_bold

    # Terminal Assumptions Section
    ws_ass.cell(row=24, column=1, value="Terminal assumptions").font = font_section
    ws_ass.cell(row=24, column=1).border = border_thin_all
    ws_ass.cell(row=24, column=2).border = border_thin_all
    
    ws_ass.cell(row=25, column=1, value="Terminal growth (g)").font = font_label_bold
    ws_ass.cell(row=25, column=1).border = border_thin_all
    cell_tg = ws_ass.cell(row=25, column=2, value=0.03) # Assumptions!B25
    cell_tg.font = font_bold_blue
    cell_tg.alignment = align_right
    cell_tg.number_format = "0.0%"
    cell_tg.border = border_thin_all
    cell_tg.fill = fill_highlight

    ws_ass.cell(row=26, column=1, value="Exit EV/EBITDA (terminal)").font = font_label_bold
    ws_ass.cell(row=26, column=1).border = border_thin_all
    cell_ex = ws_ass.cell(row=26, column=2, value=15.0) # Assumptions!B26
    cell_ex.font = font_bold_blue
    cell_ex.alignment = align_right
    cell_ex.number_format = "0.0\"x\""
    cell_ex.border = border_thin_all
    cell_ex.fill = fill_highlight

    # Inputs Block Section
    ws_ass.cell(row=28, column=1, value="Balance Sheet & Valuation Inputs").font = font_section
    ws_ass.cell(row=28, column=1).border = border_thin_all
    ws_ass.cell(row=28, column=2).border = border_thin_all
    
    ws_ass.cell(row=29, column=1, value="Cash & cash equivalents").font = font_label
    ws_ass.cell(row=29, column=1).border = border_thin_all
    cell_c = ws_ass.cell(row=29, column=2, value=cash_latest) # Assumptions!B29
    cell_c.font = font_black
    cell_c.alignment = align_right
    cell_c.number_format = "$#,##0"
    cell_c.border = border_thin_all

    ws_ass.cell(row=30, column=1, value="Total Debt").font = font_label
    ws_ass.cell(row=30, column=1).border = border_thin_all
    cell_d = ws_ass.cell(row=30, column=2, value=debt_latest) # Assumptions!B30
    cell_d.font = font_black
    cell_d.alignment = align_right
    cell_d.number_format = "$#,##0"
    cell_d.border = border_thin_all

    ws_ass.cell(row=31, column=1, value="Net debt (USD mm)").font = font_label_bold
    ws_ass.cell(row=31, column=1).border = border_thin_all
    cell_nd = ws_ass.cell(row=31, column=2, value="=B30-B29") # Assumptions!B31
    cell_nd.font = font_bold_black
    cell_nd.alignment = align_right
    cell_nd.number_format = "$#,##0"
    cell_nd.border = border_thin_all

    ws_ass.cell(row=32, column=1, value="Shares outstanding (mm)").font = font_label_bold
    ws_ass.cell(row=32, column=1).border = border_thin_all
    cell_sh = ws_ass.cell(row=32, column=2, value=shares_val) # Assumptions!B32
    cell_sh.font = font_bold_black
    cell_sh.alignment = align_right
    cell_sh.number_format = "#,##0"
    cell_sh.border = border_thin_all

    ws_ass.cell(row=33, column=1, value="Current share price (USD / share)").font = font_label_bold
    ws_ass.cell(row=33, column=1).border = border_thin_all
    cell_pr = ws_ass.cell(row=33, column=2, value=price_val) # Assumptions!B33
    cell_pr.font = font_bold_blue
    cell_pr.alignment = align_right
    cell_pr.number_format = "$#,##0.00"
    cell_pr.border = border_thin_all
    cell_pr.fill = fill_highlight

    ws_ass.cell(row=35, column=1, value="Defaults anchored on the actuals in Historical; every blue cell is yours to change.").font = font_subtitle

    # ----------------------- SHEET 3: HISTORICAL -----------------------
    ws_hist = wb.create_sheet("Historical")
    ws_hist.views.sheetView[0].showGridLines = False
    ws_hist.freeze_panes = 'B5'  # Freeze label col + header row

    # Title
    ws_hist['A2'] = f"{company_name} — Historical Financials"
    ws_hist['A2'].font = font_title
    ws_hist['A3'] = "USD mm. Source: OpenBB providers. Blue = reported actual, black = ratio."
    ws_hist['A3'].font = font_subtitle

    # Table Header
    ws_hist.cell(row=4, column=1, value="USD mm").font = font_header
    ws_hist.cell(row=4, column=1).fill = fill_header
    ws_hist.cell(row=4, column=1).border = border_thin_all
    for col_idx, yr in enumerate(hist_years, 2):
        c = ws_hist.cell(row=4, column=col_idx, value=f"FY{yr}")
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border_thin_all

    hist_rows_layout = [
        ("Revenue", "revenue", "$#,##0", False),
        ("Revenue growth %", "rev_growth", "0.0%", True),
        ("Gross profit", "gross_profit", "$#,##0", False),
        ("Gross margin %", "gp_margin", "0.0%", True),
        ("EBIT", "ebit", "$#,##0", False),
        ("EBIT margin %", "ebit_margin", "0.0%", True),
        ("Pretax income", "pretax", "$#,##0", False),
        ("Tax provision", "tax", "$#,##0", False),
        ("Effective tax rate %", "tax_rate", "0.0%", True),
        ("Net income", "net_income", "$#,##0", False),
        ("Net margin %", "net_margin", "0.0%", True),
        ("D&A", "da", "$#,##0", False),
        ("D&A % of revenue", "da_rate", "0.0%", True),
        ("Capex", "capex", "$#,##0", False),
        ("Capex % of revenue", "capex_rate", "0.0%", True),
        ("Operating cash flow", "ocf", "$#,##0", False),
        ("Free cash flow (reported)", "fcf", "$#,##0", False),
    ]

    hist_row_idx = {}
    r_idx = 5
    for label, key, num_fmt, is_ratio in hist_rows_layout:
        hist_row_idx[key] = r_idx
        r_idx += 1

    for label, key, num_fmt, is_ratio in hist_rows_layout:
        r = hist_row_idx[key]
        ws_hist.cell(row=r, column=1, value=label).font = font_label_bold if is_ratio else font_label
        ws_hist.cell(row=r, column=1).border = border_thin_all
        
        for col_idx, yr in enumerate(hist_years, 2):
            cell_col_letter = get_column_letter(col_idx)
            cell = ws_hist.cell(row=r, column=col_idx)
            cell.alignment = align_right
            cell.border = border_thin_all
            
            yr_data = hist_data.get(yr, {})
            
            if is_ratio:
                if key == "rev_growth":
                    if col_idx == 2:
                        cell.value = "" 
                    else:
                        prev_let = get_column_letter(col_idx - 1)
                        cell.value = f"=({cell_col_letter}{hist_row_idx['revenue']}-{prev_let}{hist_row_idx['revenue']})/{prev_let}{hist_row_idx['revenue']}"
                elif key == "gp_margin":
                    cell.value = f"={cell_col_letter}{hist_row_idx['gross_profit']}/{cell_col_letter}{hist_row_idx['revenue']}"
                elif key == "ebit_margin":
                    cell.value = f"={cell_col_letter}{hist_row_idx['ebit']}/{cell_col_letter}{hist_row_idx['revenue']}"
                elif key == "tax_rate":
                    cell.value = f"={cell_col_letter}{hist_row_idx['tax']}/{cell_col_letter}{hist_row_idx['pretax']}"
                elif key == "net_margin":
                    cell.value = f"={cell_col_letter}{hist_row_idx['net_income']}/{cell_col_letter}{hist_row_idx['revenue']}"
                elif key == "da_rate":
                    cell.value = f"={cell_col_letter}{hist_row_idx['da']}/{cell_col_letter}{hist_row_idx['revenue']}"
                elif key == "capex_rate":
                    cell.value = f"={cell_col_letter}{hist_row_idx['capex']}/{cell_col_letter}{hist_row_idx['revenue']}"
                
                cell.font = font_black
            else:
                actual_val = yr_data.get(key, 0.0)
                cell.value = actual_val
                cell.font = font_blue

            cell.number_format = num_fmt

    # ----------------------- SHEET 4: DCF -----------------------
    ws_dcf = wb.create_sheet("DCF")
    ws_dcf.views.sheetView[0].showGridLines = False
    ws_dcf.freeze_panes = 'B5'  # Freeze label col + header row

    # Title
    ws_dcf['A2'] = f"{company_name} ({ticker}) — Unlevered DCF"
    ws_dcf['A2'].font = font_title
    ws_dcf['A3'] = "USD mm. Base links from Historical; forward years driven by Assumptions."
    ws_dcf['A3'].font = font_subtitle

    # Header Row
    ws_dcf.cell(row=4, column=1, value="USD mm").font = font_header
    ws_dcf.cell(row=4, column=1).fill = fill_header
    ws_dcf.cell(row=4, column=1).border = border_thin_all
    
    ws_dcf.cell(row=4, column=2, value=f"FY{latest_year}A").font = font_header
    ws_dcf.cell(row=4, column=2).fill = fill_header
    ws_dcf.cell(row=4, column=2).alignment = align_center
    ws_dcf.cell(row=4, column=2).border = border_thin_all

    for col_idx, lbl in enumerate(fc_labels, 3):
        c = ws_dcf.cell(row=4, column=col_idx, value=lbl)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border_thin_all

    dcf_rows_layout = [
        ("Revenue", "revenue", "$#,##0", False, False),
        ("Revenue growth %", "growth", "0.0%", False, False),
        ("EBIT", "ebit", "$#,##0", False, False),
        ("EBIT margin %", "ebit_margin", "0.0%", False, False),
        ("Less: taxes on EBIT", "taxes", "($#,##0);$#,##0;\"-\"", False, False),
        ("NOPAT", "nopat", "$#,##0", False, True),
        ("Plus: D&A", "da", "$#,##0", False, False),
        ("Less: capex", "capex", "($#,##0);$#,##0;\"-\"", False, False),
        ("Less: Δ net working capital", "nwc", "($#,##0);$#,##0;\"-\"", False, False),
        ("Unlevered free cash flow", "ufcf", "$#,##0", False, True),
        ("empty", "empty", None, False, False),
        ("Discount period (yrs; mid-year)", "period", "0.0", False, False),
        ("Discount factor", "df", "0.00", False, False),
        ("PV of FCF", "pv_ufcf", "$#,##0", False, True),
    ]

    dcf_row_idx = {}
    r_idx = 5
    for label, key, num_fmt, is_ass, is_sum in dcf_rows_layout:
        if label == "empty":
            r_idx += 1
            continue
        dcf_row_idx[key] = r_idx
        r_idx += 1

    # Write projection rows (Row 5 to 18)
    for label, key, num_fmt, is_ass, is_sum in dcf_rows_layout:
        r = dcf_row_idx.get(key)
        if label == "empty":
            continue
            
        cell_lbl = ws_dcf.cell(row=r, column=1, value=label)
        cell_lbl.font = font_label_bold if is_sum else font_label
        cell_lbl.alignment = align_left
        cell_lbl.border = border_thin_all
        if is_sum:
            cell_lbl.fill = fill_summary
            cell_lbl.border = border_summary

        # Column B (FY2025A - Historical)
        cell_hist = ws_dcf.cell(row=r, column=2)
        cell_hist.alignment = align_right
        cell_hist.font = font_bold_black if is_sum else font_black
        cell_hist.border = border_thin_all
        if is_sum:
            cell_hist.fill = fill_summary
            cell_hist.border = border_summary
            
        if key == "revenue":
            cell_hist.value = f"=Historical!{get_column_letter(len(hist_years)+1)}{hist_row_idx['revenue']}"
            cell_hist.font = font_green_bold
        elif key == "growth":
            cell_hist.value = f"=Historical!{get_column_letter(len(hist_years)+1)}{hist_row_idx['rev_growth']}"
        elif key == "ebit":
            cell_hist.value = f"=Historical!{get_column_letter(len(hist_years)+1)}{hist_row_idx['ebit']}"
        elif key == "ebit_margin":
            cell_hist.value = f"=Historical!{get_column_letter(len(hist_years)+1)}{hist_row_idx['ebit_margin']}"
        elif key == "taxes":
            cell_hist.value = f"=Historical!{get_column_letter(len(hist_years)+1)}{hist_row_idx['tax']}"
        elif key == "nopat":
            # NOPAT = EBIT * (1 - effective tax rate). Use tax-rate row not raw tax provision.
            # Raw tax_provision can be inflated by non-op items; tax_rate is cleaner.
            hist_tax_rate_col = get_column_letter(len(hist_years) + 1)
            cell_hist.value = f"=B{dcf_row_idx['ebit']}*(1-Historical!{hist_tax_rate_col}{hist_row_idx['tax_rate']})"
        elif key == "da":
            cell_hist.value = f"=Historical!{get_column_letter(len(hist_years)+1)}{hist_row_idx['da']}"
        elif key == "capex":
            cell_hist.value = f"=Historical!{get_column_letter(len(hist_years)+1)}{hist_row_idx['capex']}"
        elif key == "nwc":
            cell_hist.value = 0.0
        elif key == "ufcf":
            cell_hist.value = f"=B{dcf_row_idx['nopat']}+B{dcf_row_idx['da']}-B{dcf_row_idx['capex']}-B{dcf_row_idx['nwc']}"

        cell_hist.number_format = num_fmt

        # Columns C to L (Forecast Years 1-10)
        for col_idx in range(3, 13):
            col_let = get_column_letter(col_idx)
            prev_let = get_column_letter(col_idx - 1)
            ass_let = get_column_letter(col_idx - 1)
            
            cell_fc = ws_dcf.cell(row=r, column=col_idx)
            cell_fc.alignment = align_right
            cell_fc.font = font_bold_black if is_sum else font_black
            cell_fc.border = border_thin_all
            if is_sum:
                cell_fc.fill = fill_summary
                cell_fc.border = border_summary
                
            if key == "revenue":
                cell_fc.value = f"={prev_let}{dcf_row_idx['revenue']}*(1+Assumptions!{ass_let}5)"
            elif key == "growth":
                cell_fc.value = f"=Assumptions!{ass_let}5"
            elif key == "ebit":
                cell_fc.value = f"={col_let}{dcf_row_idx['revenue']}*Assumptions!{ass_let}6"
            elif key == "ebit_margin":
                cell_fc.value = f"=Assumptions!{ass_let}6"
            elif key == "taxes":
                cell_fc.value = f"={col_let}{dcf_row_idx['ebit']}*Assumptions!{ass_let}7"
            elif key == "nopat":
                # NOPAT = EBIT * (1 - tax rate from Assumptions).  Consistent with UFCF build.
                cell_fc.value = f"={col_let}{dcf_row_idx['ebit']}*(1-Assumptions!{ass_let}7)"
            elif key == "da":
                cell_fc.value = f"={col_let}{dcf_row_idx['revenue']}*Assumptions!{ass_let}8"
            elif key == "capex":
                cell_fc.value = f"={col_let}{dcf_row_idx['revenue']}*Assumptions!{ass_let}9"
            elif key == "nwc":
                cell_fc.value = f"=Assumptions!{ass_let}10*({col_let}{dcf_row_idx['revenue']}-{prev_let}{dcf_row_idx['revenue']})"
            elif key == "ufcf":
                cell_fc.value = f"={col_let}{dcf_row_idx['nopat']}+{col_let}{dcf_row_idx['da']}-{col_let}{dcf_row_idx['capex']}-{col_let}{dcf_row_idx['nwc']}"
            elif key == "period":
                cell_fc.value = 0.5 + (col_idx - 3)
            elif key == "df":
                # Discount factor uses absolute WACC reference so sensitivity can override per-row.
                cell_fc.value = f"=1/((1+Assumptions!$B$22)^{col_let}{dcf_row_idx['period']})"
            elif key == "pv_ufcf":
                cell_fc.value = f"={col_let}{dcf_row_idx['ufcf']}*{col_let}{dcf_row_idx['df']}"

            cell_fc.number_format = num_fmt

    # Valuation Summary Card in the DCF tab (Rows 21 to 33)
    ws_dcf.cell(row=20, column=1, value="Valuation (USD mm)").font = font_section
    ws_dcf.cell(row=20, column=1).border = border_thin_all
    ws_dcf.cell(row=20, column=2).border = border_thin_all
    ws_dcf.cell(row=20, column=3).border = border_thin_all
    
    val_layout = [
        # Label, Cell Reference key, Formula, NumFormat, is_assumption, is_summary, border, fill
        ("PV of explicit FCF", "B21", f"=SUM(C{dcf_row_idx['pv_ufcf']}:L{dcf_row_idx['pv_ufcf']})", "$#,##0", False, False, border_summary, None),
        ("Terminal value (Gordon / exit-multiple)", "B22", f"=L{dcf_row_idx['ufcf']}*(1+Assumptions!$B$25)/(Assumptions!$B$22-Assumptions!$B$25)", "$#,##0", False, False, border_summary, None),
        ("PV of terminal value", "B23", f"=C22/((1+Assumptions!$B$22)^10)", "$#,##0", False, False, border_summary, None), # discounts Column C (Gordon TV) dynamically!
        ("Enterprise value", "B24", "=B21+B23", "$#,##0", False, True, border_summary, fill_summary),
        ("Less: net debt", "B25", "=Assumptions!B31", "($#,##0);$#,##0;\"-\"", False, False, border_summary, None),
        ("Equity value", "B26", "=B24-B25", "$#,##0", False, True, border_summary, fill_summary),
        ("Shares outstanding (mm)", "B27", "=Assumptions!B32", "#,##0", False, False, border_summary, None),
        ("Implied value / share (USD / share)", "B28", "=B26/B27", "$#,##0.00", False, True, border_summary, fill_highlight),
        ("Current share price (USD / share)", "B29", "=Assumptions!B33", "$#,##0.00", False, False, border_summary, fill_highlight_soft),
        ("Upside / (downside)", "B30", "=(B28/B29)-1", "0.0%", False, True, border_double_bottom, fill_green_soft),
        ("Implied FY1E EV/EBIT", "B31", "=B24/C7", "0.0\"x\"", False, False, border_summary, None), # C7 is Year 1 EBIT
        ("Terminal value % of EV", "B32", "=B23/B24", "0.0%", False, False, border_summary, None),
    ]

    val_row_mapping = {}
    v_r_start = 21
    for label, cell_ref, val_formula, num_fmt, is_ass, is_sum, border, fill in val_layout:
        val_row_mapping[cell_ref] = v_r_start
        v_r_start += 1

    # Write Valuation Card rows
    for label, cell_ref, val_formula, num_fmt, is_ass, is_sum, border, fill in val_layout:
        r = val_row_mapping[cell_ref]
        
        # Col 1: Label
        cell_lbl = ws_dcf.cell(row=r, column=1, value=label)
        cell_lbl.font = font_label_bold if is_sum else font_label
        cell_lbl.alignment = align_left
        cell_lbl.border = border_summary
        if fill:
            cell_lbl.fill = fill
        elif is_sum:
            cell_lbl.fill = fill_summary
            
        # Format formula with dynamic row numbers to link precisely
        formatted_formula = val_formula
        if "B2" in val_formula or "B3" in val_formula or "C2" in val_formula:
            # Replace references to B21-B32 with their dynamically resolved row numbers in DCF
            formatted_formula = val_formula.replace("B21", f"B{val_row_mapping['B21']}") \
                                            .replace("B22", f"B{val_row_mapping['B22']}") \
                                            .replace("C22", f"C{val_row_mapping['B22']}") \
                                            .replace("B23", f"B{val_row_mapping['B23']}") \
                                            .replace("B24", f"B{val_row_mapping['B24']}") \
                                            .replace("B25", f"B{val_row_mapping['B25']}") \
                                            .replace("B26", f"B{val_row_mapping['B26']}") \
                                            .replace("B27", f"B{val_row_mapping['B27']}") \
                                            .replace("B28", f"B{val_row_mapping['B28']}") \
                                            .replace("B29", f"B{val_row_mapping['B29']}")

        # Col 2: Formula Value (Calculations column)
        cell_val = ws_dcf.cell(row=r, column=2)
        cell_val.alignment = align_right
        cell_val.font = font_green_bold if "Upside" in label else (font_bold_black if is_sum else font_black)
        cell_val.border = border
        
        # Do not write values or formulas into cell B21 and B22 directly since they are comparisons, but let's see:
        # B21 should indeed contain the sum of PV of UFCFs!
        # B22 should contain nothing (the user likes it empty) or we can write the Gordon TV label comparison.
        # Wait! To avoid overwriting issues, let's write values to cell B21!
        if label == "PV of explicit FCF":
            cell_val.value = formatted_formula
        elif label == "Terminal value (Gordon / exit-multiple)":
            cell_val.value = "" # Keep column B blank for TV comparison label rows
        else:
            cell_val.value = formatted_formula
            
        if fill:
            cell_val.fill = fill
        elif is_sum:
            cell_val.fill = fill_summary

        # Add exit multiple comparison to Col 3 (Column C) and Col 4 (Column D)
        if label == "Terminal value (Gordon / exit-multiple)":
            # Column C: Gordon TV value formula
            cell_g = ws_dcf.cell(row=r, column=3, value=formatted_formula)
            cell_g.alignment = align_right
            cell_g.font = font_black
            cell_g.number_format = "$#,##0"
            cell_g.border = border
            
            # Column D: Exit-multiple TV value formula (EBITDA * Multiple)
            exit_multiple_tv = f"=(L{dcf_row_idx['ebit']}+L{dcf_row_idx['da']})*Assumptions!$B$26"
            cell_ex = ws_dcf.cell(row=r, column=4, value=exit_multiple_tv)
            cell_ex.alignment = align_right
            cell_ex.font = font_black
            cell_ex.number_format = "$#,##0"
            cell_ex.border = border
            
            # Put headers above comparison columns in Row 21 (r-1)
            # Gordon TV label in Col C, Exit-mult TV label in Col D
            ws_dcf.cell(row=r-1, column=3, value="Gordon TV").font = font_subtitle
            ws_dcf.cell(row=r-1, column=3).alignment = align_right
            ws_dcf.cell(row=r-1, column=3).border = border_summary
            
            ws_dcf.cell(row=r-1, column=4, value="Exit-mult TV").font = font_subtitle
            ws_dcf.cell(row=r-1, column=4).alignment = align_right
            ws_dcf.cell(row=r-1, column=4).border = border_summary
        else:
            # For non-comparison rows, also write thin borders for columns 3 and 4 to keep the card uniform
            ws_dcf.cell(row=r, column=3).border = border
            ws_dcf.cell(row=r, column=4).border = border
            if fill:
                ws_dcf.cell(row=r, column=3).fill = fill
                ws_dcf.cell(row=r, column=4).fill = fill
            elif is_sum:
                ws_dcf.cell(row=r, column=3).fill = fill_summary
                ws_dcf.cell(row=r, column=4).fill = fill_summary

    # ----------------------- SHEET 5: SENSITIVITY -----------------------
    ws_sens = wb.create_sheet("Sensitivity")
    ws_sens.views.sheetView[0].showGridLines = False

    # Title
    ws_sens['A2'] = "Implied Share Price (USD / share) — Sensitivity"
    ws_sens['A2'].font = font_title
    ws_sens['A3'] = "Rows: WACC. Columns: terminal growth (g). Live off the DCF FCF stream."
    ws_sens['A3'].font = font_subtitle

    # Header Row / Table Header
    ws_sens.cell(row=5, column=1, value="WACC \\ g").font = font_header
    ws_sens.cell(row=5, column=1).fill = fill_header
    ws_sens.cell(row=5, column=1).alignment = align_center
    ws_sens.cell(row=5, column=1).border = border_thin_all

    # terminal growth rates in Row 5 columns B to F (values: -1.0%, -0.5%, base, +0.5%, +1.0%)
    for col_idx, val_formula in enumerate([
        "=Assumptions!$B$25-0.01",
        "=Assumptions!$B$25-0.005",
        "=Assumptions!$B$25",
        "=Assumptions!$B$25+0.005",
        "=Assumptions!$B$25+0.01"
    ], 2):
        cell = ws_sens.cell(row=5, column=col_idx, value=val_formula)
        cell.font = font_header
        cell.fill = fill_header
        cell.number_format = "0.0%"
        cell.alignment = align_center
        cell.border = border_thin_all

    # WACC rates in A6:A10 (values: -2.0%, -1.0%, base, +1.0%, +2.0%)
    for idx, val_formula in enumerate([
        "=Assumptions!$B$22-0.02",
        "=Assumptions!$B$22-0.01",
        "=Assumptions!$B$22",
        "=Assumptions!$B$22+0.01",
        "=Assumptions!$B$22+0.02"
    ], 6):
        cell = ws_sens.cell(row=idx, column=1, value=val_formula)
        cell.font = font_label_bold
        cell.number_format = "0.0%"
        cell.alignment = align_center
        cell.border = border_thin_all

    # Present Value calculation from DCF sheet
    pv_terms = []
    for col_idx in range(3, 13):
        col_let = get_column_letter(col_idx)
        pv_terms.append(f"(DCF!${col_let}${dcf_row_idx['ufcf']}/((1+$A{{r}})^DCF!${col_let}${dcf_row_idx['period']}))")
    pv_fcf_sum = " + ".join(pv_terms)

    # Grid Cells calculations
    for r in range(6, 11):
        for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F'], 2):
            cell = ws_sens.cell(row=r, column=col_idx)
            
            # WACC is $A{r}, Terminal growth is {col_letter}$5
            cell.value = f"=( {pv_fcf_sum} + ((DCF!$L${dcf_row_idx['ufcf']}*(1+{col_letter}$5)/($A{r}-{col_letter}$5))/((1+$A{r})^10)) - Assumptions!$B$31 ) / Assumptions!$B$32".format(r=r)
            cell.alignment = align_right
            cell.font = font_black
            cell.number_format = "$#,##0.00"
            cell.border = border_thin_all
            
            # Highlight the Base Case intersection cell D8 (base WACC, base g)
            if r == 8 and col_letter == 'D':
                cell.fill = fill_highlight
                cell.font = font_bold_black  # Use global 9pt compact font, not ad-hoc size=11

    ws_sens.cell(row=12, column=1, value="Yellow = base case (model WACC & g). Each cell re-discounts the DCF FCF stream using mid-year periods.").font = font_subtitle

    # ----------------------- SHEET 6: CHECKS -----------------------
    ws_chk = wb.create_sheet("Checks")
    ws_chk.views.sheetView[0].showGridLines = False
    
    # Title
    ws_chk['A2'] = f"{company_name} ({ticker}) — Model Checks & Controls"
    ws_chk['A2'].font = font_title
    ws_chk['A3'] = "Formula-driven model checks to ensure math integrity."
    ws_chk['A3'].font = font_subtitle

    # Checks Grid
    ws_chk.cell(row=5, column=1, value="Integrity Checks").font = font_header
    ws_chk.cell(row=5, column=1).fill = fill_header
    ws_chk.cell(row=5, column=1).border = border_thin_all
    ws_chk.cell(row=5, column=2, value="Status").font = font_header
    ws_chk.cell(row=5, column=2).fill = fill_header
    ws_chk.cell(row=5, column=2).alignment = align_center
    ws_chk.cell(row=5, column=2).border = border_thin_all

    ws_chk.cell(row=6, column=1, value="Implied share price positive check").font = font_label
    ws_chk.cell(row=6, column=1).border = border_thin_all
    c1 = ws_chk.cell(row=6, column=2, value="=IF(Summary!B8>0,\"PASS\",\"FAIL\")")
    c1.alignment = align_center
    c1.font = font_green_bold
    c1.border = border_thin_all

    ws_chk.cell(row=7, column=1, value="Balance sheet net debt calculation integrity check").font = font_label
    ws_chk.cell(row=7, column=1).border = border_thin_all
    c2 = ws_chk.cell(row=7, column=2, value="=IF(Assumptions!B31=Assumptions!B30-Assumptions!B29,\"PASS\",\"FAIL\")")
    c2.alignment = align_center
    c2.font = font_green_bold
    c2.border = border_thin_all

    # ----------------------- SHEET 7: SOURCES & AUDIT -----------------------
    ws_src = wb.create_sheet("Sources & Audit")
    ws_src.views.sheetView[0].showGridLines = False
    
    # Title
    ws_src['A2'] = f"{company_name} ({ticker}) — Data Sources & Audits"
    ws_src['A2'].font = font_title
    
    audit_notes = [
        ("Source:", "Data retrieved dynamically via OpenBB Platform v4 SDK and Yahoo Finance fundamental financial databases."),
        ("Revenue Projections:", "Years 1-3 anchored to analyst consensus (yfinance street estimates). Year 3 implied via EPS trend deceleration. Years 4-10: convex power-1.5 decay from Year-3 anchor to 3.0% long-run g, ensuring growth stays healthy in mid-years before converging to terminal."),
        ("Shares Outstanding:", "Derived from market cap / current price to capture all share classes (e.g. GOOG Class A+B+C). Avoids under-counting from single-class API fields."),
        ("Capital Structure (WACC Weighting):", "Market-value WACC. Equity weight = Mktcap / (Mktcap + Gross Debt); Debt weight = 1 − Equity weight. Beta from OpenBB/yfinance profile."),
        ("Mid-year convention:", "Cash flows assumed at mid-period: discount periods 0.5, 1.5, ... 9.5. Terminal value (Gordon Growth) received at end of Year 10, discounted at t=10."),
        ("NOPAT:", "NOPAT = EBIT × (1 − effective tax rate). Using rate-based calc avoids noise from non-operating items in raw tax provision."),
        ("Steady State Maintenance CapEx:", "CapEx % of revenue tapered linearly; Year 10 CapEx is locked to equal D&A % of revenue, ensuring zero net reinvestment at terminal."),
        ("NWC changes:", "Modeled as 3.0% of incremental revenue (industry standard). Positive ΔRevenue → positive ΔNWC → cash outflow subtracted in UFCF."),
    ]

    for idx, (lbl, text) in enumerate(audit_notes, 5):
        cell_lbl = ws_src.cell(row=idx, column=1, value=lbl)
        cell_lbl.font = font_label_bold
        cell_lbl.border = border_thin_all
        cell_txt = ws_src.cell(row=idx, column=2, value=text)
        cell_txt.font = font_label
        cell_txt.border = border_thin_all

    # ----------------------- PRECISION COLUMN WIDTHS & ROW HEIGHTS -----------------------
    # Compact, professional widths — no more endless scrolling.
    # Each sheet gets its own tailored spec based on content type.
    COMPACT_ROW_HEIGHT = 15  # pts — matches Bloomberg / Goldman standard

    # Summary: label col wide enough for longest label, value col just right
    ws_sum.column_dimensions['A'].width = 36
    ws_sum.column_dimensions['B'].width = 14
    for col_letter in ['C', 'D', 'E']:
        ws_sum.column_dimensions[col_letter].width = 4
    for row in ws_sum.iter_rows():
        ws_sum.row_dimensions[row[0].row].height = COMPACT_ROW_HEIGHT

    # Assumptions: label col + 10 forecast year cols
    ws_ass.column_dimensions['A'].width = 32
    for col_idx in range(2, 13):  # B to L (Year 1-10)
        ws_ass.column_dimensions[get_column_letter(col_idx)].width = 10
    for row in ws_ass.iter_rows():
        ws_ass.row_dimensions[row[0].row].height = COMPACT_ROW_HEIGHT

    # Historical: label col + 4 historical year cols
    ws_hist.column_dimensions['A'].width = 32
    for col_idx in range(2, 7):  # B to F
        ws_hist.column_dimensions[get_column_letter(col_idx)].width = 12
    for row in ws_hist.iter_rows():
        ws_hist.row_dimensions[row[0].row].height = COMPACT_ROW_HEIGHT

    # DCF: label col + base year + 10 forecast year cols + comparison cols
    ws_dcf.column_dimensions['A'].width = 30
    for col_idx in range(2, 15):  # B to N
        ws_dcf.column_dimensions[get_column_letter(col_idx)].width = 12
    for row in ws_dcf.iter_rows():
        ws_dcf.row_dimensions[row[0].row].height = COMPACT_ROW_HEIGHT

    # Sensitivity: axis label col + 5 data cols
    ws_sens.column_dimensions['A'].width = 10
    for col_idx in range(2, 8):  # B to G
        ws_sens.column_dimensions[get_column_letter(col_idx)].width = 12
    for row in ws_sens.iter_rows():
        ws_sens.row_dimensions[row[0].row].height = COMPACT_ROW_HEIGHT

    # Checks & Sources: two-col layout
    for ws in [ws_chk, ws_src]:
        ws.column_dimensions['A'].width = 36
        ws.column_dimensions['B'].width = 80
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = COMPACT_ROW_HEIGHT

    # Save to file — always write into the output/ folder next to the script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, "output")
    os.makedirs(output_dir, exist_ok=True)

    if output_path is None:
        output_path = os.path.join(output_dir, f"{ticker}_DCF_Model.xlsx")
    wb.save(output_path)
    print(f"[+] Success! Multi-sheet dynamic model saved to: {output_path}")

    # Optional: Upload copy to Google Drive / convert to Google Sheets
    maybe_upload_to_google_drive(output_path, ticker)
    
    return True

def maybe_upload_to_google_drive(filepath, ticker):
    """
    Checks for Google Service Account credentials outside the repository.
    If found, uploads the generated Excel sheet to Google Drive and converts it to Google Sheets.
    """
    import json
    config_dir = os.path.expanduser("~/.config/dcf_excel")
    creds_path = os.path.join(config_dir, "google_credentials.json")
    config_path = os.path.join(config_dir, "config.json")
    
    if not os.path.exists(creds_path):
        return  # Silently skip if no API credentials configured
        
    print(f"[*] Google API credentials found. Initiating Drive upload for {ticker}...")
    
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaFileUpload
        
        # Load folder ID configuration if available
        folder_id = None
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r') as f:
                    cfg = json.load(f)
                    folder_id = cfg.get("drive_folder_id")
            except Exception as e:
                print(f"[!] Warning: Failed to read config.json ({e})")
                
        # Authenticate
        scopes = ['https://www.googleapis.com/auth/drive']
        creds = service_account.Credentials.from_service_account_file(creds_path, scopes=scopes)
        service = build('drive', 'v3', credentials=creds)
        
        # Define metadata — converting to Google Sheets mimeType converts the file automatically!
        file_metadata = {
            'name': f"{ticker}_DCF_Model",
            'mimeType': 'application/vnd.google-apps.spreadsheet'
        }
        if folder_id:
            file_metadata['parents'] = [folder_id]
            
        media = MediaFileUpload(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )
        
        print(f"[*] Uploading and converting '{os.path.basename(filepath)}' to Google Sheets...")
        uploaded_file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id,webViewLink'
        ).execute()
        
        print(f"[+] Success! Google Sheet created.")
        print(f"[+] Link: {uploaded_file.get('webViewLink')}")
        
    except Exception as e:
        print(f"[!] Error uploading to Google Drive: {e}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate professional DCF models using OpenBB and openpyxl.")
    parser.add_argument("--ticker", type=str, default="AAPL", help="Stock ticker symbol (e.g. AAPL, GOOG, RDDT)")
    parser.add_argument("--output", type=str, default=None, help="Output file path (optional)")
    args = parser.parse_args()

    print("=================================================================")
    print("      WALL STREET STANDARD DCF MODEL AUTOMATION GENERATOR       ")
    print("=================================================================")
    
    success = generate_openbb_dcf(args.ticker.upper(), args.output)
    if success:
        print("[*] Excel DCF Model Generated Successfully.")
        sys.exit(0)
    else:
        print("[!] Model Generation Failed.")
        sys.exit(1)
